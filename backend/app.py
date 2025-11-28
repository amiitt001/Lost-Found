import sys
import os
import json
from io import BytesIO
try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    # Defer a helpful error until runtime when a file is uploaded
    load_workbook = None
from flask import Flask, request, jsonify
from flask_cors import CORS  # Needed to allow the frontend to talk to the backend
# --- Configuration ---
app = Flask(__name__)
# Enable CORS for all routes for frontend access
CORS(app)
# Simple in-memory storage for raw uploaded data
DATA_STORE = {}
DATA_ID_COUNTER = 0

# ==========================================
# Section A: Helper Utilities (from utils.py)
# ==========================================

def clean_header(header):
    """Normalizes header string."""
    return str(header).strip().lower() if header else ""


def find_col_index(headers, candidates):
    """Finds the 0-based index of a column header by trying multiple candidates."""
    normalized_headers = [clean_header(h) for h in headers]
    for candidate in candidates:
        cand = candidate.lower()
        if cand in normalized_headers:
            return normalized_headers.index(cand)
    return -1


def split_roll_branch(s):
    """Splits student string into roll and branch."""
    if not s:
        return {"roll": "", "branch": ""}
    s = str(s).strip()
    parts = s.split()
    roll = parts[0]
    branch = " ".join(parts[1:]) if len(parts) > 1 else ""
    return {"roll": roll, "branch": branch}


# ==========================================
# Section B: Core Seating Logic (from logic.py)
# ==========================================

def generate_seating_plan(students, rooms, pattern='standard'):
    """Allocates students to rooms based on the specific pattern."""

    student_queue = []
    for pair in students:
        if pair.get('s1'):
            student_queue.append({'val': pair['s1'], 'orig': 'Series 1', 'id': pair['id']})
        if pair.get('s2'):
            student_queue.append({'val': pair['s2'], 'orig': 'Series 2', 'id': pair['id']})

    processed_rooms = [r.copy() for r in rooms]
    queue_idx = 0
    total_students = len(student_queue)

    for room in processed_rooms:
        rows, cols = room['rows'], room['cols']
        coords = []  # (row_index, col_index, capacity)

        # --- Pattern Logic Switch ---
        if pattern == 'columnar':
            for c in range(cols):
                for r in range(rows):
                    coords.append((r, c, 2))
        elif pattern == 'snake-vertical':
            for c in range(cols):
                col_coords = [(r, c, 2) for r in range(rows)]
                if c % 2 == 1:
                    col_coords.reverse()
                coords.extend(col_coords)
        elif pattern == 'checkerboard':
            for r in range(rows):
                for c in range(cols):
                    if (r + c) % 2 == 0:
                        coords.append((r, c, 2))
        elif pattern == 'single':
            for r in range(rows):
                for c in range(cols):
                    coords.append((r, c, 1))
        elif pattern == 'alternate-rows':
            for r in range(rows):
                capacity = 2 if r % 2 == 0 else 1
                for c in range(cols):
                    coords.append((r, c, capacity))
        elif pattern == 'hybrid':
            for c in range(cols):
                capacity = 2 if c % 2 == 0 else 1
                for r in range(rows):
                    coords.append((r, c, capacity))
        else:  # Standard (Z) & Snake (S)
            for r in range(rows):
                row_coords = [(r, c, 2) for c in range(cols)]
                if pattern == 'snake' and r % 2 == 1:
                    row_coords.reverse()
                coords.extend(row_coords)

        # --- Fill Grid ---
        grid = [[{'left': None, 'right': None} for _ in range(cols)] for _ in range(rows)]
        assignments_count = 0

        for r, c, cap in coords:
            if queue_idx < total_students:
                s1_obj = student_queue[queue_idx]
                grid[r][c]['left'] = s1_obj
                queue_idx += 1
                assignments_count += 1

                if cap == 2 and queue_idx < total_students:
                    s2_obj = student_queue[queue_idx]
                    grid[r][c]['right'] = s2_obj
                    queue_idx += 1
                    assignments_count += 1

        room['grid'] = grid
        room['assigned_count'] = assignments_count

    unallocated = student_queue[queue_idx:]

    return processed_rooms, unallocated

@app.route('/create-paymen', methods=['POST'])
def create_payment():
    """Handles payment creation."""
    try:
        data = request.json
        amount = data.get('amount')
        currency = data.get('currency')
        description = data.get('description')
        customer_email = data.get('customerEmail')

        # Validate amount and currency
        if not amount or not currency:
            return jsonify({"error": "Amount and currency are required"}), 400

        # Create payment transaction
        payment_response = create_payment_transaction(amount, currency, description, customer_email)

        return jsonify(payment_response), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles Excel file upload, parses data, and stores it."""
    global DATA_ID_COUNTER

    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file and file.filename.endswith(('.xlsx', '.xls')):
        # Provide a clear runtime error if openpyxl isn't available
        if load_workbook is None:
            return jsonify({
                "error": "Server missing dependency 'openpyxl'. Please install it: pip install openpyxl"
            }), 500
        try:
            # Read file into memory buffer once and open workbook
            file_bytes = file.read()
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]

            # Identify Columns
            idx_roll1 = find_col_index(headers, ["roll no. series-1", "series-1", "roll1"])
            idx_roll2 = find_col_index(headers, ["roll no. series-2", "series-2", "roll2"])
            idx_room = find_col_index(headers, ["room no.", "room"])
            idx_rows = find_col_index(headers, ["rows", "row", "no. of rows"])
            idx_cols = find_col_index(headers, ["columns", "cols"])
            idx_college = find_col_index(headers, ["college name", "college"])
            idx_exam = find_col_index(headers, ["exam name", "exam"])

            if idx_roll1 == -1 or idx_roll2 == -1 or idx_room == -1:
                return jsonify({"error": "Missing required columns: Series-1, Series-2, or Room No."}), 400

            students = []
            rooms_map = {}
            total_students = 0

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):

                # Student pairs
                s1 = str(row[idx_roll1]).strip() if idx_roll1 < len(row) and row[idx_roll1] else ""
                s2 = str(row[idx_roll2]).strip() if idx_roll2 < len(row) and row[idx_roll2] else ""

                if s1 or s2:
                    students.append({'s1': s1, 's2': s2, 'id': i + 2})
                    if s1:
                        total_students += 1
                    if s2:
                        total_students += 1

                # Room configurations (must be unique by name)
                r_name = str(row[idx_room]).strip() if idx_room < len(row) and row[idx_room] else ""
                if r_name and r_name not in rooms_map:
                    try:
                        r_rows = int(row[idx_rows]) if idx_rows != -1 and idx_rows < len(row) and row[idx_rows] else 0
                        r_cols = int(row[idx_cols]) if idx_cols != -1 and idx_cols < len(row) and row[idx_cols] else 0
                    except (ValueError, TypeError):
                        r_rows, r_cols = 0, 0

                    if r_rows > 0 and r_cols > 0:
                        rooms_map[r_name] = {
                            'name': r_name, 'rows': r_rows, 'cols': r_cols,
                            'college': str(row[idx_college]) if idx_college != -1 and idx_college < len(row) and row[idx_college] else "",
                            'exam': str(row[idx_exam]) if idx_exam != -1 and idx_exam < len(row) and row[idx_exam] else ""
                        }

            if not students:
                return jsonify({"error": "No student records found."}), 400
            if not rooms_map:
                return jsonify({"error": "No valid room configurations found (check Row/Col counts)."}), 400

            DATA_ID_COUNTER += 1
            data_id = str(DATA_ID_COUNTER)
            DATA_STORE[data_id] = {
                'students': students,
                'rooms': list(rooms_map.values()),
                'total_students': total_students
            }

            return jsonify({
                "message": "File processed successfully",
                "data_id": data_id,
                "total_students": total_students,
                "room_count": len(rooms_map)
            }), 200

        except Exception as e:
            print(f"Server error during processing: {e}")
            return jsonify({"error": f"An error occurred during file processing: {str(e)}"}), 500

    return jsonify({"error": "Invalid file type. Please upload XLSX or XLS."}), 400


@app.route('/calculate', methods=['POST'])
def calculate_seating():
    """Handles the request to run the seating algorithm based on pattern."""
    data = request.get_json()
    data_id = data.get('data_id')
    pattern = data.get('pattern', 'standard')

    if data_id not in DATA_STORE:
        return jsonify({"error": "Data ID not found. Please upload the file again."}), 404

    raw_data = DATA_STORE[data_id]

    try:
        processed_rooms, unallocated = generate_seating_plan(
            raw_data['students'],
            raw_data['rooms'],
            pattern
        )

        # Prepare unallocated list for frontend (splitting roll/branch for display)
        unallocated_for_display = []
        for student in unallocated:
            details = split_roll_branch(student['val'])
            unallocated_for_display.append({
                'roll': details['roll'],
                'branch': details['branch'],
                'orig': student.get('orig', '')
            })

        # Prepare room grids for display (splitting roll/branch for DeskCard)
        for room in processed_rooms:
            for r in range(room['rows']):
                for c in range(room['cols']):
                    desk = room['grid'][r][c]
                    if desk and desk.get('left'):
                        left_obj = desk['left']
                        left_details = split_roll_branch(left_obj.get('val'))
                        left_details['orig'] = left_obj.get('orig', '')
                        desk['left'] = left_details
                    if desk and desk.get('right'):
                        right_obj = desk['right']
                        right_details = split_roll_branch(right_obj.get('val'))
                        right_details['orig'] = right_obj.get('orig', '')
                        desk['right'] = right_details

        return jsonify({
            "rooms": processed_rooms,
            "unallocated": unallocated_for_display,
            "total_students": raw_data['total_students']
        }), 200

    except Exception as e:
        print(f"Calculation error: {e}")
        return jsonify({"error": "Calculation failed. Check room and student data consistency."}), 500


if __name__ == '__main__':
    # Flask runs on port 5000 by default
    print("Starting Flask server on http://127.0.0.1:5000")
    # Set host='0.0.0.0' to be accessible from outside the container/localhost
    app.run(host='0.0.0.0', port=5000, debug=True)
